import openpyxl
from django.contrib import admin
from django.http import HttpResponse
from django.utils import timezone

from core.models import Metric

# Register your models here.
admin.site.register(Metric)


class AbstractAdmin(admin.ModelAdmin):
    def get_obj_id(self, obj):
        return f"✏️ {obj.id}"

    @admin.action(description='Обновить с вызовом сигнала')
    def call_update_signal(self, request, queryset):
        for query in queryset:
            query.updated_at = timezone.now()
            query.save()

    @admin.action(description='Экспорт в xlsx')
    def export_to_xlsx(self, request, queryset):
        wb = openpyxl.Workbook()
        default_sheet = wb.active
        ws = wb.create_sheet("Данные")
        wb.remove_sheet(default_sheet)
        file_name = f"{queryset.model.__name__}.xlsx"
        model_fields = queryset.model._meta.get_fields()
        for index, field in enumerate(model_fields):
            if hasattr(field, 'verbose_name') and field.verbose_name:
                field_name = field.verbose_name
            else:
                field_name = field.name
            ws.cell(row=1, column=index + 1).value = field_name
        for row_index, query in enumerate(queryset, start=2):
            for column_index, field in enumerate(model_fields, start=1):
                if field.is_relation:
                    try:
                        field_value = list(getattr(query, field.name).all().values_list('id', flat=True))
                    except AttributeError:
                        try:
                            field_value = list(getattr(query, f"{field.name}_set").all().values_list('id', flat=True))
                        except AttributeError:
                            field_value = None
                else:
                    field_value = getattr(query, field.name)

                if field_value:
                    field_value = str(field_value)
                else:
                    field_value = "-"
                ws.cell(row=row_index, column=column_index).value = field_value

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': f'attachment; '
                                       f'filename="{file_name}"'},
        )
        wb.save(response)

        return response

        # ws["A1"] = f"[{self.application_instance.id}] {self.application_instance.name}"
        # ws["A2"] = f"Список посетителей на {now.strftime('%d.%m.%Y %H:%M')}"
        # for column, field in enumerate(self.application_instance.visitors_additional_fields.all().order_by('order'),
        #                                start=1):
        #     ws.cell(row=3, column=column).value = f"{field.name}"

    readonly_fields = ['id', 'created_at', 'updated_at']
    actions = [call_update_signal, export_to_xlsx]
